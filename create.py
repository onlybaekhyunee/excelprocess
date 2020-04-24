import openpyxl
import datetime
wb = openpyxl.Workbook()
ws = wb.active
print(ws.title)
ws['A1'] = '520'
ws.append(['yfc','36','UESTC'])
ws['A3'] = datetime.datetime.now()
wb.save('demo.xlsx')
